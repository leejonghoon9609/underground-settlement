# redeploy trigger 2026-06-06-v2-1530
# force vercel rebuild
from http.server import BaseHTTPRequestHandler
import json
import base64
import os
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment
import copy
import re

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template.xlsx')

RATES = {
    'realtime': {'labor': 6412, 'machine': 9746},
    'probe':    {'labor': 3206, 'machine': 4873},
}
RATIOS = {
    'indirectLabor': 0.10, 'accident': 0.0356, 'employment': 0.0101,
    'pension': 0.0475, 'health': 0.03595,
    'elderly': 0.1314, 'genMgmt': 0.03, 'profit': 0.135,
    'contract': 0.749, 'vat': 0.10,
}

# 사업구분 순서 고정 (value값 그대로 사용)
TANGO_ORDER = [
    '[A-2. 인입관로] 인입 관로 공급',
    '[B-3. 기간선로] 신설/증설/보강',
    '[B-3. 기간선로] 신축국사 연계 간선선로',
    '[C-2. 프론트홀 선로(5G)] 용량증설(5G)',
    '[E-4. 지장이설] 원인자 공사',
    '[E-4. 지장이설] 지중 인프라 확보',
    '[E-4. 지장이설] 순수 지장 이설',
    '[G-3. 프론트홀 선로(4G)] 용량증설(4G)',
    '[O-1. 취약개선] 중요 선로구간 Risk 해소(지상고 불량 등)',
]

def calc_cost(exposed_km, probe_km, method, survey_name=''):
    probe_m = probe_km * 1000
    safety_rate = 0.0178 if '수도권지사' in (survey_name or '') else 0.0164
    r = RATES[method]
    dl = int(probe_m * r['labor'])
    me = int(probe_m * r['machine'])
    il = int(dl * RATIOS['indirectLabor'])
    lt = dl + il
    ac = int(lt * RATIOS['accident'])
    em = int(lt * RATIOS['employment'])
    pe = int(dl * RATIOS['pension'])
    he = int(dl * RATIOS['health'])
    sa = int(dl * safety_rate)
    el = int(he * RATIOS['elderly'])
    et = ac + em + pe + he + sa + el + me
    gm = int((lt + et) * RATIOS['genMgmt'])
    pr = int((lt + et + gm) * RATIOS['profit'])
    wc = lt + et + gm + pr
    ct = int((wc - sa) * RATIOS['contract']) + sa
    fi = (ct // 1000) * 1000
    vt = round(fi * RATIOS['vat'])
    return {
        'directLabor': dl, 'indirectLabor': il, 'laborTotal': lt,
        'machineExp': me, 'accident': ac, 'employment': em,
        'pension': pe, 'health': he, 'safety': sa, 'elderly': el,
        'expTotal': et, 'generalMgmt': gm, 'profit': pr,
        'workCost': wc, 'finalCost': fi, 'vat': vt, 'totalWithVat': fi + vt,
    }

COST_ROW_MAP = {
    10: 'directLabor', 15: 'indirectLabor', 16: 'laborTotal',
    19: 'accident', 20: 'employment', 21: 'pension',
    22: 'health', 23: 'safety', 24: 'elderly',
    27: 'machineExp', 28: 'expTotal', 29: 'generalMgmt',
    30: 'profit', 31: 'workCost', 32: 'finalCost',
    40: 'finalCost', 41: 'vat', 42: 'totalWithVat',
}

def get_fmt(cell):
    return {
        'font': copy.copy(cell.font),
        'alignment': copy.copy(cell.alignment),
        'border': copy.copy(cell.border),
        'fill': copy.copy(cell.fill),
        'number_format': cell.number_format,
    }

def apply_fmt(cell, fmt):
    cell.font = fmt['font']
    cell.alignment = fmt['alignment']
    cell.border = fmt['border']
    cell.fill = fmt['fill']
    cell.number_format = fmt['number_format']

# 원가계산서 사업 블록(3열 × 5~42행)을 마스터(13~15열)에서 대상 열로 복제
def clone_cost_block(ws, src_start, dst_start, rows=range(5, 43)):
    from openpyxl.formula.translate import Translator
    for r in rows:
        for k in range(3):
            src = ws.cell(r, src_start + k)
            dst = ws.cell(r, dst_start + k)
            dst.font          = copy.copy(src.font)
            dst.alignment     = copy.copy(src.alignment)
            dst.border        = copy.copy(src.border)
            dst.fill          = copy.copy(src.fill)
            dst.number_format = src.number_format
            v = src.value
            if isinstance(v, str) and v.startswith('='):
                # 수식은 열 이동만큼 자동 이동 (절대참조 $E$41 등은 그대로 유지)
                dst.value = Translator(v, origin=src.coordinate).translate_formula(dst.coordinate)
            else:
                dst.value = v
    # 제목행(5행) 병합 재생성
    # (열 너비는 복제하지 않음 - 기존 블록2~16과 동일한 기본 너비 유지, 인쇄 폭 불변)
    ws.merge_cells(start_row=5, start_column=dst_start, end_row=5, end_column=dst_start + 2)

class handler(BaseHTTPRequestHandler):

    def send_cors_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, Accept')

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_cors_headers()
        self.send_header('Content-Length', '0')
        self.end_headers()

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body = json.loads(self.rfile.read(length))
            projects = body['projects']

            wb = openpyxl.load_workbook(TEMPLATE_PATH)

            # ── 공공측량 갑지 ──────────────────────────────
            ws_gap = wb['공공측량 갑지']

            # 제목(A1) 생성
            year  = body.get('year', '2026')
            month = body.get('month', '')
            branch = ''
            for p in projects:
                name = p.get('surveyName', '')
                m = re.search(r'\(([^)]+지사)\)', name)
                if m:
                    branch = m.group(1)
                    break
            # 연도 뒤 2자리 (2026 → 26), 월 앞 0 제거 (06 → 6)
            yy = str(year)[-2:]
            mm = str(month).lstrip('0') or str(month)
            ws_gap.cell(1, 1).value = f"{yy}년 기성준공내역서 {mm}월_도급_SKTNS_{branch}_측량(대원항업)"

            # 서식 기준: 원본 4행, 12행
            data_fmt = [get_fmt(ws_gap.cell(4, col)) for col in range(1, 10)]
            sum_fmt  = [get_fmt(ws_gap.cell(12, col)) for col in range(1, 10)]

            # 4~20행 완전 초기화 (값 + 서식 + 테두리 모두)
            from openpyxl.styles import PatternFill, Border, Side
            empty_fill = PatternFill(fill_type=None)
            empty_border = Border(
                left=Side(style=None), right=Side(style=None),
                top=Side(style=None), bottom=Side(style=None)
            )
            for row in range(4, 21):
                for col in range(1, 10):
                    cell = ws_gap.cell(row, col)
                    cell.value = None
                    cell.fill = empty_fill
                    cell.border = empty_border
            # 사업구분별 집계 (tangoType 전체 라벨 기준)
            from collections import defaultdict
            summary = defaultdict(lambda: {'exposed': 0.0, 'probe': 0.0, 'tango': 0.0, 'cost': 0})
            for p in projects:
                t = p.get('tangoType', '')
                summary[t]['exposed'] += float(p.get('exposedKm', 0))
                summary[t]['probe']   += float(p.get('probeKm', 0))
                summary[t]['tango']   += float(p.get('tangoKm', 0))
                summary[t]['cost']    += int(p.get('finalCost', 0))

            # 데이터 있는 항목만 순서대로 (TANGO_ORDER 기준)
            active = [label for label in TANGO_ORDER if label in summary]
            # 목록에 없는 사업구분도 누락되지 않도록 뒤에 추가 (안전장치)
            for label in summary:
                if label and label not in active:
                    active.append(label)

            # 데이터행 쓰기 (4행부터)
            tot = {'survey': 0.0, 'tango': 0.0, 'exposed': 0.0, 'probe': 0.0, 'sub': 0.0, 'cost': 0}
            for i, label in enumerate(active):
                row = 4 + i
                s = summary[label]
                survey = round(s['exposed'] + s['probe'], 3)
                sub    = round(s['exposed'] + s['probe'], 3)

                for col in range(1, 10):
                    apply_fmt(ws_gap.cell(row, col), data_fmt[col - 1])

                ws_gap.cell(row, 1).value = label
                ws_gap.cell(row, 2).value = round(survey, 3)
                ws_gap.cell(row, 3).value = round(s['tango'], 3)
                ws_gap.cell(row, 4).value = round(s['exposed'], 3)
                ws_gap.cell(row, 5).value = round(s['probe'], 3)
                ws_gap.cell(row, 6).value = round(sub, 3)
                ws_gap.cell(row, 7).value = s['cost']

                tot['survey']  += survey
                tot['tango']   += s['tango']
                tot['exposed'] += s['exposed']
                tot['probe']   += s['probe']
                tot['sub']     += sub
                tot['cost']    += s['cost']

            # 합계행
            sum_row = 4 + len(active)
            for col in range(1, 10):
                apply_fmt(ws_gap.cell(sum_row, col), sum_fmt[col - 1])

            ws_gap.cell(sum_row, 1).value = '합          계'
            ws_gap.cell(sum_row, 1).alignment = Alignment(horizontal='center', vertical='center')
            ws_gap.cell(sum_row, 2).value = round(tot['survey'], 3)
            ws_gap.cell(sum_row, 3).value = round(tot['tango'], 3)
            ws_gap.cell(sum_row, 4).value = round(tot['exposed'], 3)
            ws_gap.cell(sum_row, 5).value = round(tot['probe'], 3)
            ws_gap.cell(sum_row, 6).value = round(tot['sub'], 3)
            ws_gap.cell(sum_row, 7).value = tot['cost']

            # ── 세부내역 시트 ──────────────────────────────
            ws_detail = wb['세부내역']
            for i, p in enumerate(projects):
                row = 5 + i
                if row > 20:
                    break
                cost = calc_cost(
                    float(p.get('exposedKm', 0)),
                    float(p.get('probeKm', 0)),
                    p.get('method', 'probe'),
                    p.get('surveyName', '')
                )
                ws_detail.cell(row, 3).value  = p.get('gubun', '')
                ws_detail.cell(row, 4).value  = p.get('region', '')
                ws_detail.cell(row, 5).value  = p.get('surveyName', '')
                ws_detail.cell(row, 6).value  = p.get('workCode', '')
                ws_detail.cell(row, 7).value  = p.get('workName', '')
                ws_detail.cell(row, 8).value  = p.get('tangoType', '')
                ws_detail.cell(row, 10).value = float(p.get('tangoKm', 0))
                ws_detail.cell(row, 11).value = float(p.get('exposedKm', 0))
                ws_detail.cell(row, 12).value = float(p.get('probeKm', 0))
                ws_detail.cell(row, 15).value = cost['finalCost']
                ws_detail.cell(row, 16).value = p.get('remark', '')

            # 사업 수만큼만 남기고 빈 데이터행 삭제 → 소계가 바로 뒤로 붙음
            from openpyxl.utils import get_column_letter as _gcl
            n_detail = min(len(projects), 16)   # 데이터행 5~20 (최대 16개)
            sub_row = 5 + n_detail              # 소계가 위치할 최종 행
            if n_detail < 16:
                # 소계 라벨 병합(B21:D21)은 delete_rows가 안 옮기므로 수동 처리
                for m in list(ws_detail.merged_cells.ranges):
                    if m.min_row == 21 and m.min_col == 2 and m.max_col == 4:
                        ws_detail.unmerge_cells(str(m))
                ws_detail.delete_rows(5 + n_detail, 16 - n_detail)
                ws_detail.merge_cells(start_row=sub_row, start_column=2,
                                      end_row=sub_row, end_column=4)
            # 소계 SUM 범위를 실제 데이터 범위로 재설정 (delete_rows가 수식은 안 고침)
            last_row = 4 + n_detail
            for col in (9, 10, 11, 12, 13, 15):   # I,J,K,L,M,O
                L = _gcl(col)
                ws_detail.cell(sub_row, col).value = f'=SUM({L}5:{L}{last_row})'

            # ── 원가계산서 시트 ────────────────────────────
            ws_cost = wb['원가계산서']

            # 안전관리비 비율 동적 반영 (23행 E열)
            safety_rate = 0.0178 if any('수도권지사' in p.get('surveyName', '') for p in projects) else 0.0164
            ws_cost.cell(23, 5).value = safety_rate

            for i, p in enumerate(projects):
                cost = calc_cost(
                    float(p.get('exposedKm', 0)),
                    float(p.get('probeKm', 0)),
                    p.get('method', 'probe'),
                    p.get('surveyName', '')
                )
                start_col = 13 + i * 3
                # 첫 블록(13열)은 템플릿에 이미 있고, 그 다음부터는 마스터 블록 복제
                if i > 0:
                    clone_cost_block(ws_cost, 13, start_col)
                ws_cost.cell(5, start_col).value = f"{i+1}. {p.get('workCode', '')}"
                for row, key in COST_ROW_MAP.items():
                    val = cost.get(key, 0)
                    ws_cost.cell(row, start_col).value     = val
                    ws_cost.cell(row, start_col + 1).value = 0
                    ws_cost.cell(row, start_col + 2).value = val

            # ── 사업별 캡쳐용 탭 (맨 뒤에 순서대로) ──────────
            # 작업자가 캡쳐 이미지를 붙여넣을 빈 탭. 페이지 설정만 동일하게 맞춤.
            for i, p in enumerate(projects):
                sheet_name = f"{i+1}. {p.get('workCode', '')}"
                sheet_name = sheet_name[:31]  # 엑셀 시트명 31자 제한
                ws_cap = wb.create_sheet(title=sheet_name)  # 맨 뒤에 추가됨
                ws_cap.page_setup.paperSize   = 9          # A4
                ws_cap.page_setup.orientation = 'portrait'  # 세로
                ws_cap.page_setup.scale       = 75          # 확대/축소 75%
                ws_cap.page_margins.top    = 0.75
                ws_cap.page_margins.bottom = 0.75
                ws_cap.page_margins.left   = 0.7
                ws_cap.page_margins.right  = 0.7
                ws_cap.page_margins.header = 0.3
                ws_cap.page_margins.footer = 0.3
                ws_cap.print_area = 'A1:X29'
                ws_cap.sheet_view.view = 'pageBreakPreview'  # 페이지 나누기 미리 보기로 열기

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            result_b64 = base64.b64encode(output.read()).decode('utf-8')

            resp = json.dumps({'file': result_b64}).encode('utf-8')
            self.send_response(200)
            self.send_cors_headers()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', str(len(resp)))
            self.end_headers()
            self.wfile.write(resp)

        except Exception as e:
            import traceback
            err_msg = traceback.format_exc()
            resp = json.dumps({'error': str(e), 'traceback': err_msg}).encode('utf-8')
            self.send_response(500)
            self.send_cors_headers()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', str(len(resp)))
            self.end_headers()
            self.wfile.write(resp)
